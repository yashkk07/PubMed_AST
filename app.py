import streamlit as st
import requests
import urllib.parse
import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np
import math
import datetime
import time
import re
import os
import io
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from collections import OrderedDict
from wordcloud import WordCloud
import nltk
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from io import BytesIO
import logging
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

# Configure logging to file instead of displaying on screen
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                   filename='pubmed_app.log',
                   filemode='a')
logger = logging.getLogger('pubmed_app')

# Suppress warnings in streamlit
# st.set_option('deprecation.showPyplotGlobalUse', False)

# Download necessary NLTK data silently without user-visible warnings
try:
    nltk.download('punkt', quiet=True)
    nltk.download('stopwords', quiet=True)
    from nltk.tokenize import word_tokenize
    from nltk.corpus import stopwords
    NLTK_AVAILABLE = True
    logger.info("NLTK resources loaded successfully")
except Exception as e:
    NLTK_AVAILABLE = False
    logger.error(f"NLTK error: {str(e)}")

# Set Streamlit page configuration
st.set_page_config(
    page_title="PubMed Research Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# App title and styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E88E5;
        text-align: center;
        margin-bottom: 1rem;
    }
    .subheader {
        font-size: 1.5rem;
        color: #0D47A1;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .section-header {
        font-size: 1.25rem;
        color: #1565C0;
        margin-top: 1.5rem;
        margin-bottom: 0.75rem;
    }
    .insight-box {
        background-color: #1A237E;
        color: white;
        border-radius: 5px;
        padding: 1rem;
        margin-bottom: 1rem;
    }
    .text-centered {
        text-align: center;
    }
    .footer {
        text-align: center;
        margin-top: 3rem;
        padding-top: 1rem;
        border-top: 1px solid #1565C0;
        color: #BBDEFB;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #0D47A1;
        color: white;
        border-radius: 4px 4px 0px 0px;
        gap: 1px;
        padding-top: 10px;
        padding-bottom: 10px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1976D2;
        color: white;
    }
    
    /* Main background and text color */
    .stApp {
        background-color: #121212;
        color: #FFFFFF;
    }
    
    /* Make download buttons stand out */
    .stDownloadButton button {
        background-color: #1976D2 !important;
        color: white !important;
        font-weight: bold !important;
        border: 1px solid #1565C0 !important;
        padding: 10px 20px !important;
    }
    
    .stDownloadButton button:hover {
        background-color: #1565C0 !important;
        border-color: #0D47A1 !important;
    }
    
    /* Style dataframes for dark mode */
    .stDataFrame {
        background-color: #212121;
    }
    
    /* Style form elements */
    [data-testid="stForm"] {
        border-color: #424242 !important;
        background-color: #212121;
        padding: 20px;
    }
    
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #1A1A1A;
    }
    
    /* Hide streamlit warnings/errors */
    [data-testid="stSidebar"] [data-testid="stException"] {
        display: none;
    }
    div[data-testid="stStatusWidget"] {
        display: none;
    }
</style>
<h1 class="main-header">PubMed Research Analyzer</h1>
<p class="text-centered">Extract, analyze, and download PubMed articles before and after FDA approval.</p>
""", unsafe_allow_html=True)

# ------------------ PubMed Retrieval Functions ------------------

DB = 'pubmed'
BASEURL_SRCH = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi'
BASEURL_FTCH = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi'
BATCH_NUM = 1000  # records per batch

def mkquery(base_url, params):
    """Build a URL with proper URL encoding."""
    query = '&'.join([f'{key}={urllib.parse.quote(str(value))}' for key, value in params.items()])
    return f"{base_url}?{query}"

def getXmlFromURL(base_url, params):
    """Return the XML ElementTree for a given URL built from parameters."""
    url = mkquery(base_url, params)
    response = requests.get(url)
    return ET.fromstring(response.text)

def getTextFromNode(root, path, default=""):
    """Safely extract text from an XML node specified by an XPath."""
    node = root.find(path)
    return node.text.strip() if node is not None and node.text is not None else default

def parse_month(month_str):
    """
    Convert month string to numeric month value (1-12).
    Handles numeric strings, full month names, and abbreviations.
    """
    if not month_str:
        return "1"  # Default to January if no month
    
    # If already numeric
    if month_str.isdigit():
        month_num = int(month_str)
        if 1 <= month_num <= 12:
            return str(month_num)
    
    # If it's a text month name
    month_str = month_str.strip().lower()
    month_map = {
        'jan': '1', 'january': '1',
        'feb': '2', 'february': '2',
        'mar': '3', 'march': '3',
        'apr': '4', 'april': '4',
        'may': '5',
        'jun': '6', 'june': '6',
        'jul': '7', 'july': '7',
        'aug': '8', 'august': '8',
        'sep': '9', 'september': '9', 'sept': '9',
        'oct': '10', 'october': '10',
        'nov': '11', 'november': '11',
        'dec': '12', 'december': '12'
    }
    
    for abbr, num in month_map.items():
        if month_str.startswith(abbr):
            return num
    
    # If we couldn't determine the month, default to January
    return "1"

# ------------------ Processing Each Article ------------------

def process_article(article):
    """
    Given an XML element for a PubMed article, extract basic information and author details
    with full publication date precision. Also extracts DOI and creates article link.
    """
    record = OrderedDict()
    record["PMID"] = getTextFromNode(article, "./MedlineCitation/PMID", "")
    record["JournalTitle"] = getTextFromNode(article, "./MedlineCitation/Article/Journal/Title", "")
    record["ArticleTitle"] = getTextFromNode(article, "./MedlineCitation/Article/ArticleTitle", "")
    
    # Extract DOI
    doi = ""
    # Try to get DOI from first ArticleId with DOI type
    article_ids = article.findall(".//ArticleId")
    for article_id in article_ids:
        if article_id.get("IdType") == "doi" and article_id.text:
            doi = article_id.text.strip()
            break
    
    # If not found, try getting DOI from ELocation element
    if not doi:
        elocation_ids = article.findall(".//ELocationID")
        for eloc in elocation_ids:
            if eloc.get("EIdType") == "doi" and eloc.text:
                doi = eloc.text.strip()
                break
    
    record["DOI"] = doi
    
    # Create article link
    pmid = record["PMID"]
    if pmid:
        record["ArticleLink"] = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
    else:
        record["ArticleLink"] = ""
    
    # Extract full publication date information - first try ArticleDate
    pub_year = getTextFromNode(article, "./MedlineCitation/Article/ArticleDate/Year", "")
    pub_month = getTextFromNode(article, "./MedlineCitation/Article/ArticleDate/Month", "")
    pub_day = getTextFromNode(article, "./MedlineCitation/Article/ArticleDate/Day", "")
    
    # If ArticleDate is not available, try JournalIssue/PubDate
    if not pub_year:
        pub_year = getTextFromNode(article, "./MedlineCitation/Article/Journal/JournalIssue/PubDate/Year", "")
        pub_month = getTextFromNode(article, "./MedlineCitation/Article/Journal/JournalIssue/PubDate/Month", "")
        pub_day = getTextFromNode(article, "./MedlineCitation/Article/Journal/JournalIssue/PubDate/Day", "")
    
    # If still no month, try MedlineDate which might contain a string like "2014 Mar-Apr"
    if not pub_month:
        medline_date = getTextFromNode(article, "./MedlineCitation/Article/Journal/JournalIssue/PubDate/MedlineDate", "")
        if medline_date:
            # Try to extract year and month from strings like "2014 Mar-Apr" or "2023 Jan"
            parts = medline_date.split()
            if len(parts) >= 2 and parts[0].isdigit():
                pub_year = parts[0]
                month_part = parts[1].split('-')[0]  # Take first month if range
                pub_month = parse_month(month_part)
    
    # If year is missing or invalid, try to extract from other fields
    if not pub_year:
        pub_date = getTextFromNode(article, "./MedlineCitation/Article/Journal/JournalIssue/PubDate/MedlineDate", "")
        if pub_date:
            # Try to extract just the year from any date string
            year_match = re.search(r'\b(19|20)\d{2}\b', pub_date)
            if year_match:
                pub_year = year_match.group(0)
    
    # Clean and convert months to standardized format
    pub_month = parse_month(pub_month)
    
    # Default day to 1 if not available
    if not pub_day or not pub_day.isdigit():
        pub_day = "1"
    
    # Store full date components
    record["PublicationYear"] = pub_year
    record["PublicationMonth"] = pub_month
    record["PublicationDay"] = pub_day
    
    # Create a full date string in format "YYYY-MM-DD"
    try:
        year = int(pub_year) if pub_year else 0
        month = int(pub_month) if pub_month else 1
        day = int(pub_day) if pub_day else 1
        
        if year > 0:
            # Ensure valid month (1-12)
            month = max(1, min(12, month))
            # Ensure valid day (1-31 depending on month)
            max_days = [31, 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28, 
                       31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            day = max(1, min(max_days[month-1], day))
            
            # Format the date
            record["PublicationDate"] = f"{year:04d}-{month:02d}-{day:02d}"
        else:
            record["PublicationDate"] = ""
    except (ValueError, TypeError):
        # If any conversion fails, store empty string
        record["PublicationDate"] = ""
    
    abstracts = article.findall("./MedlineCitation/Article/Abstract/AbstractText")
    if abstracts:
        record["Abstract"] = " ".join([a.text.strip() for a in abstracts if a.text])
    else:
        record["Abstract"] = ""
    
    authors = article.findall("./MedlineCitation/Article/AuthorList/Author")
    author_count = 1
    
    for auth in authors:
        name = ""
        collective = auth.find("CollectiveName")
        if collective is not None and collective.text:
            name = collective.text.strip()
        else:
            fore = auth.find("ForeName")
            last = auth.find("LastName")
            if fore is not None and last is not None:
                name = f"{fore.text.strip()} {last.text.strip()}"
            elif fore is not None:
                name = fore.text.strip()
            elif last is not None:
                name = last.text.strip()
        
        aff_elem = auth.find("./AffiliationInfo/Affiliation")
        affiliation = aff_elem.text.strip() if (aff_elem is not None and aff_elem.text) else ""
        
        record[f"Author{author_count}"] = name
        record[f"Affiliation{author_count}"] = affiliation
        author_count += 1
    
    return record

def fetch_pubmed_articles(query, batch_size=BATCH_NUM, limit=None, progress_callback=None):
    """
    Use ESearch and EFetch to retrieve PubMed articles matching a query.
    Returns a list of OrderedDict records (one per article), each processed by process_article().
    Optional limit parameter to restrict the number of articles processed.
    """
    params = {"db": DB, "term": query, "usehistory": "y", "retmax": 0}
    root = getXmlFromURL(BASEURL_SRCH, params)
    count = int(getTextFromNode(root, "./Count", "0"))
    
    if count == 0:
        return []
    
    # If limit is specified, adjust count
    if limit and limit < count:
        count = limit
    
    query_key = getTextFromNode(root, "./QueryKey", "")
    webenv = getTextFromNode(root, "./WebEnv", "")
    
    records = []
    iterCount = math.ceil(count / batch_size)
    for i in range(iterCount):
        params_fetch = {
            "db": DB,
            "query_key": query_key,
            "WebEnv": webenv,
            "retstart": i * batch_size,
            "retmax": batch_size,
            "retmode": "xml"
        }
        
        # For the last batch with a limit, adjust retmax
        if limit and (i+1) * batch_size > limit:
            params_fetch["retmax"] = limit - i * batch_size
        
        root_fetch = getXmlFromURL(BASEURL_FTCH, params_fetch)
        for art in root_fetch.findall(".//PubmedArticle"):
            record = process_article(art)
            records.append(record)
            
            # Check if we've reached the limit
            if limit and len(records) >= limit:
                break
                
        # If we've reached the limit, break out of the loop
        if limit and len(records) >= limit:
            break
            
        # Update progress
        if progress_callback:
            progress_callback((i + 1) / iterCount)
            
        time.sleep(0.34)  # Respect rate limits
    return records
# ------------------ Data Verification Functions ------------------
def check_date_range(df, dataset_type, start_date, end_date):
    """
    Check if dates are within the expected range using full publication date.
    Handles different precision levels with appropriate rounding.
    
    For imprecise dates:
    - Year only: For "before" category, use Dec 31; for "after", use Jan 1
    - Year-month: For "before" category, use last day of month; for "after", use first day
    
    Returns the count of invalid dates found.
    """
    invalid_count = 0
    rows_to_drop = []
    
    # Month name to number mapping (both full names and abbreviations)
    month_map = {
        'jan': 1, 'january': 1,
        'feb': 2, 'february': 2,
        'mar': 3, 'march': 3,
        'apr': 4, 'april': 4,
        'may': 5,
        'jun': 6, 'june': 6,
        'jul': 7, 'july': 7,
        'aug': 8, 'august': 8,
        'sep': 9, 'september': 9, 'sept': 9,
        'oct': 10, 'october': 10,
        'nov': 11, 'november': 11,
        'dec': 12, 'december': 12
    }
    
    for idx, row in df.iterrows():
        # First try to use the PublicationDate field if available
        if not pd.isna(row.get('PublicationDate', None)) and row.get('PublicationDate', '') != '':
            try:
                # Try to parse the date
                pub_date = datetime.datetime.strptime(row['PublicationDate'], "%Y-%m-%d")
                
                # Check if date is outside the expected range
                if dataset_type == "before" and (pub_date < start_date or pub_date > end_date):
                    rows_to_drop.append(idx)
                    invalid_count += 1
                elif dataset_type == "after" and (pub_date < start_date):
                    rows_to_drop.append(idx)
                    invalid_count += 1
                continue  # Skip to next record if we've processed this one
            except (ValueError, TypeError):
                # If parsing fails, continue to component-based processing
                pass
        
        # If PublicationDate doesn't exist or couldn't be parsed, try component-based approach
        pub_year = row.get("PublicationYear", "")
        pub_month = row.get("PublicationMonth", "")
        pub_day = row.get("PublicationDay", "")
        
        if pub_year and str(pub_year).isdigit():
            year = int(pub_year)
            
            # Different handling based on available precision
            if pub_month and str(pub_month).isdigit():
                month = int(pub_month)
                
                if pub_day and str(pub_day).isdigit():
                    # Full date precision
                    day = int(pub_day)
                    try:
                        pub_date = datetime.datetime(year, month, day)
                    except ValueError:
                        # Invalid date (e.g., Feb 30) - use month-level precision
                        if dataset_type == "before":
                            # Get the last day of the month
                            if month == 12:
                                last_day = 31
                            else:
                                try:
                                    next_month = datetime.datetime(year, month+1, 1)
                                    last_day = (next_month - datetime.timedelta(days=1)).day
                                except ValueError:
                                    last_day = 28  # Default to a safe value
                            
                            pub_date = datetime.datetime(year, month, last_day)
                        else:  # "after"
                            pub_date = datetime.datetime(year, month, 1)
                else:
                    # Year-month precision - use last day or first day
                    if dataset_type == "before":
                        # Get the last day of the month
                        if month == 12:
                            last_day = 31
                        else:
                            try:
                                next_month = datetime.datetime(year, month+1, 1)
                                last_day = (next_month - datetime.timedelta(days=1)).day
                            except ValueError:
                                # Invalid month, default to day 1
                                last_day = 1
                        
                        try:
                            pub_date = datetime.datetime(year, month, last_day)
                        except ValueError:
                            # If invalid date, default to Jan 1
                            pub_date = datetime.datetime(year, 1, 1)
                    else:  # "after"
                        try:
                            pub_date = datetime.datetime(year, month, 1)
                        except ValueError:
                            # Invalid month, default to Jan 1
                            pub_date = datetime.datetime(year, 1, 1)
            else:
                # Year-only precision - use Dec 31 or Jan 1
                if dataset_type == "before":
                    pub_date = datetime.datetime(year, 12, 31)
                else:  # "after"
                    pub_date = datetime.datetime(year, 1, 1)
            
            # Check if date is outside the expected range
            if dataset_type == "before" and (pub_date < start_date or pub_date > end_date):
                rows_to_drop.append(idx)
                invalid_count += 1
            elif dataset_type == "after" and (pub_date < start_date):
                rows_to_drop.append(idx)
                invalid_count += 1
    
    # Drop rows with invalid dates
    if rows_to_drop:
        df.drop(rows_to_drop, inplace=True)
    
    return invalid_count


def find_max_author_column(df):
    """Find the maximum author column that has non-empty values"""
    max_author = 0
    author_pattern = re.compile(r"^Author(\d+)$")
    
    # For each row, find the highest author number with data
    for _, row in df.iterrows():
        row_max = 0
        for col in df.columns:
            match = author_pattern.match(col)
            if match:
                author_num = int(match.group(1))
                # Check if this specific cell has content
                if pd.notna(row[col]) and row[col] != '':
                    if author_num > row_max:
                        row_max = author_num
        
        # Update the overall max if this row has a higher number
        if row_max > max_author:
            max_author = row_max
    
    return max_author

def trim_author_columns(df, max_author):
    """Trim excess author columns beyond the maximum needed"""
    columns_to_keep = []
    author_pattern = re.compile(r"^Author(\d+)$")
    affiliation_pattern = re.compile(r"^Affiliation(\d+)$")
    profile_pattern = re.compile(r"^Profile(\d+)$")
    
    # Add all non-author/affiliation columns
    for col in df.columns:
        if (not author_pattern.match(col) and 
            not affiliation_pattern.match(col) and 
            not profile_pattern.match(col)):
            columns_to_keep.append(col)
    
    # Add author/affiliation columns up to max_author
    for i in range(1, max_author + 1):
        author_col = f"Author{i}"
        if author_col in df.columns:
            columns_to_keep.append(author_col)
        
        affiliation_col = f"Affiliation{i}"
        if affiliation_col in df.columns:
            columns_to_keep.append(affiliation_col)
        
        profile_col = f"Profile{i}"
        if profile_col in df.columns:
            columns_to_keep.append(profile_col)
    
    # Return the dataframe with only the needed columns
    return df[columns_to_keep]

def process_dataset(df, dataset_type, start_date, end_date):
    """Process a dataset to check dates and trim author columns"""
    original_row_count = len(df)
    original_column_count = len(df.columns)
    
    # Make a copy to avoid SettingWithCopyWarning
    df = df.copy()
    
    # 1. Check and filter by date if within range
    invalid_date_count = check_date_range(df, dataset_type, start_date, end_date)
    
    # 2. Trim excess author columns
    max_author_column = find_max_author_column(df)
    df = trim_author_columns(df, max_author_column)
    
    return df, {
        "original_rows": original_row_count,
        "original_columns": original_column_count,
        "final_rows": len(df),
        "final_columns": len(df.columns),
        "invalid_dates": invalid_date_count,
        "max_author": max_author_column,
        "columns_removed": original_column_count - len(df.columns)
    }

def verify_dates_and_recategorize(records_before, records_after, fda_date):
    """
    Double-check dates and move articles to the correct category, with improved handling 
    for different date precision levels.
    
    For imprecise dates:
    - Year only: For "before" category, use Dec 31 of the year; for "after", use Jan 1
    - Year-month: For "before" category, use last day of month; for "after", use first day
    
    Returns verified before/after records lists and counts of moved articles.
    """
    verified_before = []
    verified_after = []
    moved_to_before = 0
    moved_to_after = 0
    
    # Process "before" records
    for record in records_before:
        # First try to use the PublicationDate field if available
        if record.get("PublicationDate", ""):
            try:
                pub_date = datetime.datetime.strptime(record["PublicationDate"], "%Y-%m-%d")
                if pub_date > fda_date:
                    verified_after.append(record)
                    moved_to_after += 1
                else:
                    verified_before.append(record)
                continue  # Skip to next record
            except (ValueError, TypeError):
                # If parsing fails, continue to component-based processing
                pass
        
        # Component-based approach for imprecise dates
        pub_year = record.get("PublicationYear", "")
        pub_month = record.get("PublicationMonth", "")
        pub_day = record.get("PublicationDay", "")
        
        if pub_year and str(pub_year).isdigit():
            year = int(pub_year)
            
            # Different handling based on available precision
            if pub_month and str(pub_month).isdigit():
                month = int(pub_month)
                
                if pub_day and str(pub_day).isdigit():
                    # Full date precision
                    day = int(pub_day)
                    try:
                        pub_date = datetime.datetime(year, month, day)
                    except ValueError:
                        # For "before" records with invalid dates, use conservative approach
                        # Treat as Dec 31 of the year (maximizing chance it belongs in "before")
                        pub_date = datetime.datetime(year, 12, 31)
                else:
                    # Year-month precision - for "before" records, use last day of month
                    try:
                        if month == 12:
                            last_day = 31
                        else:
                            next_month = datetime.datetime(year, month+1, 1)
                            last_day = (next_month - datetime.timedelta(days=1)).day
                        
                        pub_date = datetime.datetime(year, month, last_day)
                    except ValueError:
                        # Invalid month, default to Dec 31
                        pub_date = datetime.datetime(year, 12, 31)
            else:
                # Year-only precision - for "before" records, use Dec 31
                pub_date = datetime.datetime(year, 12, 31)
            
            # Check against FDA date
            if pub_date > fda_date:
                verified_after.append(record)
                moved_to_after += 1
            else:
                verified_before.append(record)
        else:
            # If no parseable year, keep in original category
            verified_before.append(record)
    
    # Process "after" records
    for record in records_after:
        # First try to use the PublicationDate field if available
        if record.get("PublicationDate", ""):
            try:
                pub_date = datetime.datetime.strptime(record["PublicationDate"], "%Y-%m-%d")
                if pub_date <= fda_date:
                    verified_before.append(record)
                    moved_to_before += 1
                else:
                    verified_after.append(record)
                continue  # Skip to next record
            except (ValueError, TypeError):
                # If parsing fails, continue to component-based processing
                pass
        
        # Component-based approach for imprecise dates
        pub_year = record.get("PublicationYear", "")
        pub_month = record.get("PublicationMonth", "")
        pub_day = record.get("PublicationDay", "")
        
        if pub_year and str(pub_year).isdigit():
            year = int(pub_year)
            
            # Different handling based on available precision
            if pub_month and str(pub_month).isdigit():
                month = int(pub_month)
                
                if pub_day and str(pub_day).isdigit():
                    # Full date precision
                    day = int(pub_day)
                    try:
                        pub_date = datetime.datetime(year, month, day)
                    except ValueError:
                        # For "after" records with invalid dates, use conservative approach
                        # Treat as Jan 1 of the year (maximizing chance it belongs in "after")
                        pub_date = datetime.datetime(year, 1, 1)
                else:
                    # Year-month precision - for "after" records, use first day of month
                    try:
                        pub_date = datetime.datetime(year, month, 1)
                    except ValueError:
                        # Invalid month, default to Jan 1
                        pub_date = datetime.datetime(year, 1, 1)
            else:
                # Year-only precision - for "after" records, use Jan 1
                pub_date = datetime.datetime(year, 1, 1)
            
            # Check against FDA date
            if pub_date <= fda_date:
                verified_before.append(record)
                moved_to_before += 1
            else:
                verified_after.append(record)
        else:
            # If no parseable year, keep in original category
            verified_after.append(record)
    
    return verified_before, verified_after, moved_to_before, moved_to_after

def deduplicate_records(records_before, records_after):
    # First deduplicate within each dataset
    before_pmids = {}
    after_pmids = {}
    
    # Deduplicate within before dataset
    for record in records_before:
        pmid = record.get("PMID", "")
        if pmid and pmid not in before_pmids:
            before_pmids[pmid] = record
    
    # Deduplicate within after dataset
    for record in records_after:
        pmid = record.get("PMID", "")
        if pmid and pmid not in after_pmids:
            after_pmids[pmid] = record
    
    # Convert to lists
    unique_before = list(before_pmids.values())
    unique_after = list(after_pmids.values())
    
    # Now deduplicate between datasets
    # Before records take precedence
    pmid_seen = set(before_pmids.keys())
    
    # Filter after dataset to remove any PMIDs already in before dataset
    final_after = [record for record in unique_after if record.get("PMID", "") not in pmid_seen]
    
    return unique_before, final_after

# ------------------ Data Analysis Functions ------------------

def create_top_journals_chart(df):
    """Create a bar chart of top journals by publication count"""
    if len(df) == 0 or 'JournalTitle' not in df.columns:
        return None
    
    journal_counts = df['JournalTitle'].value_counts().reset_index()
    journal_counts.columns = ['Journal', 'Count']
    
    # Get top 10 journals or all if less than 10
    top_n = min(10, len(journal_counts))
    top_journals = journal_counts.head(top_n)
    
    # Create horizontal bar chart
    fig = px.bar(
        top_journals, 
        y='Journal', 
        x='Count',
        orientation='h',
        title=f'Top {top_n} Journals by Publication Count',
        height=400,
        color='Count',
        color_continuous_scale=px.colors.sequential.Blues
    )
    
    fig.update_layout(
        xaxis_title='Number of Publications',
        yaxis_title='Journal',
        yaxis={'categoryorder':'total ascending'},
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(30,30,30,1)'
    )
    
    return fig

def create_yearly_trend_chart(df):
    """Create a line chart showing publication trends over years"""
    if len(df) == 0 or 'PublicationYear' not in df.columns:
        return None
    
    # Convert publication year to numeric and create year count
    df_year = df.copy()
    df_year['PublicationYear'] = pd.to_numeric(df_year['PublicationYear'], errors='coerce')
    df_year = df_year.dropna(subset=['PublicationYear'])
    
    if len(df_year) == 0:
        return None
    
    # Round to integer years
    df_year['PublicationYear'] = df_year['PublicationYear'].astype(int)
    
    # Count by year
    year_counts = df_year['PublicationYear'].value_counts().reset_index()
    year_counts.columns = ['Year', 'Count']
    year_counts = year_counts.sort_values('Year')
    
    # Create line chart
    fig = px.line(
        year_counts, 
        x='Year', 
        y='Count',
        markers=True,
        title='Publication Trend by Year',
        height=400
    )
    
    fig.update_layout(
        xaxis_title='Publication Year',
        yaxis_title='Number of Publications',
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(30,30,30,1)'
    )
    
    return fig

def create_monthly_trend_chart(df):
    """Create a line chart showing publication trends by month"""
    if len(df) == 0 or 'PublicationYear' not in df.columns or 'PublicationMonth' not in df.columns:
        return None
    
    # Convert publication year and month to numeric values
    df_date = df.copy()
    df_date['PublicationYear'] = pd.to_numeric(df_date['PublicationYear'], errors='coerce')
    df_date['PublicationMonth'] = pd.to_numeric(df_date['PublicationMonth'], errors='coerce')
    
    # Drop rows with missing year or month
    df_date = df_date.dropna(subset=['PublicationYear', 'PublicationMonth'])
    
    if len(df_date) == 0:
        return None
    
    # Create date strings in format "YYYY-MM" for sorting
    df_date['YearMonth'] = df_date.apply(
        lambda row: f"{int(row['PublicationYear']):04d}-{int(row['PublicationMonth']):02d}", 
        axis=1
    )
    
    # Count by year-month
    date_counts = df_date['YearMonth'].value_counts().reset_index()
    date_counts.columns = ['YearMonth', 'Count']
    date_counts = date_counts.sort_values('YearMonth')
    
    # Create line chart
    fig = px.line(
        date_counts, 
        x='YearMonth', 
        y='Count',
        markers=True,
        title='Publication Trend by Month',
        height=400
    )
    
    fig.update_layout(
        xaxis_title='Publication Month',
        yaxis_title='Number of Publications',
        xaxis={'type': 'category'},  # Treat as categorical to show all months
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(30,30,30,1)'
    )
    
    # Rotate x-axis labels for better readability
    fig.update_xaxes(tickangle=45)
    
    return fig

def create_author_collaboration_network(df, top_n=20):
    """Create a simple author collaboration network visualization"""
    if len(df) == 0:
        return None
    
    # Get all author columns
    author_cols = [col for col in df.columns if col.startswith('Author') and col != 'Author']
    if not author_cols:
        return None
    
    # Collect all authors
    all_authors = []
    for _, row in df.iterrows():
        paper_authors = [row[col] for col in author_cols if pd.notna(row[col]) and row[col] != '']
        if len(paper_authors) > 1:  # Only consider papers with multiple authors
            all_authors.extend(paper_authors)
    
    # Count author occurrences
    author_counts = pd.Series(all_authors).value_counts()
    
    # Get top authors
    top_authors = author_counts.head(top_n).index.tolist()
    
    # Create edges between co-authors
    edges = []
    for _, row in df.iterrows():
        paper_authors = [row[col] for col in author_cols if pd.notna(row[col]) and row[col] != '' and row[col] in top_authors]
        for i in range(len(paper_authors)):
            for j in range(i+1, len(paper_authors)):
                edges.append((paper_authors[i], paper_authors[j]))
    
    # Count edge occurrences
    edge_counts = pd.Series(edges).value_counts().reset_index()
    if len(edge_counts) == 0:
        return None
        
    edge_counts.columns = ['pairs', 'weight']
    edge_counts[['source', 'target']] = pd.DataFrame(edge_counts['pairs'].tolist(), index=edge_counts.index)
    
    # Create nodes dataframe with author frequencies
    nodes = pd.DataFrame({
        'name': top_authors,
        'size': [author_counts[author] for author in top_authors]
    })
    
    # Create network visualization using Plotly
    node_x = []
    node_y = []
    node_text = []
    node_size = []
    
    # Simple circular layout for nodes
    for i, (_, row) in enumerate(nodes.iterrows()):
        angle = 2 * math.pi * i / len(nodes)
        x = math.cos(angle)
        y = math.sin(angle)
        node_x.append(x)
        node_y.append(y)
        node_text.append(row['name'])
        node_size.append(row['size'] * 10)  # Scale node size
    
    # Create edges
    edge_x = []
    edge_y = []
    edge_width = []
    
    for _, row in edge_counts.iterrows():
        source_idx = nodes[nodes['name'] == row['source']].index[0]
        target_idx = nodes[nodes['name'] == row['target']].index[0]
        
        x0, y0 = node_x[source_idx], node_y[source_idx]
        x1, y1 = node_x[target_idx], node_y[target_idx]
        
        edge_x.extend([x0, x1, None])
        edge_y.extend([y0, y1, None])
        edge_width.append(row['weight'])
    
    # Create plot
    fig = go.Figure()
    
    # Add edges
    fig.add_trace(go.Scatter(
        x=edge_x, y=edge_y,
        line=dict(width=0.5, color='#888'),
        hoverinfo='none',
        mode='lines'
    ))
    
    # Add nodes
    fig.add_trace(go.Scatter(
        x=node_x, y=node_y,
        mode='markers',
        marker=dict(
            size=node_size,
            color='#1565C0',
            line=dict(width=1, color='#333')
        ),
        text=node_text,
        hoverinfo='text'
    ))
    
    fig.update_layout(
        title='Top Author Collaboration Network',
        showlegend=False,
        hovermode='closest',
        margin=dict(b=20, l=5, r=5, t=40),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        height=600,
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    return fig

def generate_wordcloud(df, column='Abstract'):
    """Generate a word cloud from text in a specified column"""
    if len(df) == 0 or column not in df.columns:
        return None
    
    try:
        # Combine all text
        text = ' '.join(df[column].dropna().astype(str))
        if not text or len(text) < 10:
            return None
        
        # Simple tokenization without NLTK
        all_words = text.lower().split()
        
        # Try to use NLTK if available
        try:
            if NLTK_AVAILABLE:
                stop_words = list(stopwords.words('english'))
                custom_stopwords = ['disease', 'patient', 'treatment', 'study', 'use', 'result', 
                                   'method', 'conclusion', 'background', 'objective']
                stop_words.extend(custom_stopwords)
                all_words = word_tokenize(text.lower())
            else:
                # Fallback stopwords
                stop_words = ['a', 'an', 'the', 'and', 'or', 'but', 'if', 'because', 'as', 'what',
                            'when', 'where', 'how', 'who', 'which', 'this', 'that', 'these', 'those',
                            'then', 'just', 'so', 'than', 'such', 'both', 'through', 'about', 'for',
                            'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had',
                            'having', 'do', 'does', 'did', 'doing', 'to', 'from', 'in', 'out', 'on',
                            'off', 'over', 'under', 'with', 'by', 'of', 'at', 'into', 'during', 'before',
                            'after', 'above', 'below', 'up', 'down', 'can', 'could', 'will', 'would',
                            'may', 'might', 'must', 'should', 'not', 'no', 'nor', 'only', 'own', 'same',
                            'too', 'very', 'disease', 'patient', 'treatment', 'study', 'use', 'result']
        except Exception:
            # If everything fails, use basic fallback
            stop_words = ['a', 'an', 'the', 'and', 'or', 'but', 'if', 'disease', 'patient']
        
        # Filter words
        filtered_words = [word for word in all_words if word.isalpha() and word not in stop_words and len(word) > 2]
        
        # If we don't have enough words after filtering, return None
        if len(filtered_words) < 10:
            logger.warning("Not enough words for wordcloud after filtering")
            return None
        
        # Create word cloud
        wordcloud = WordCloud(
            width=800, 
            height=400, 
            background_color='black', 
            max_words=100,
            contour_width=1,
            contour_color='steelblue',
            colormap='Blues'
        ).generate(' '.join(filtered_words))
        
        # Create a new figure with explicit figure object
        fig = plt.figure(figsize=(10, 5))
        plt.style.use('dark_background')
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        plt.tight_layout()
        
        return fig
    except Exception as e:
        logger.error(f"Error generating wordcloud: {str(e)}")
        return None

def generate_bigram_analysis(df, column='Abstract', top_n=15):
    """Generate bigram analysis from text"""
    if len(df) == 0 or column not in df.columns:
        return None
    
    try:
        # Combine all text
        texts = df[column].dropna().astype(str).tolist()
        if not texts:
            return None
        
        # Initialize vectorizer for bigrams
        vectorizer = CountVectorizer(
            ngram_range=(2, 2),  # bigrams
            stop_words="english",  # Use built-in English stopwords
            min_df=2  # minimum document frequency
        )
        
        # Get bigram counts
        X = vectorizer.fit_transform(texts)
        bigram_counts = np.asarray(X.sum(axis=0)).flatten()
        
        # Get bigram names and counts
        bigrams = [(word, count) for word, count in zip(vectorizer.get_feature_names_out(), bigram_counts)]
        bigrams.sort(key=lambda x: x[1], reverse=True)
        
        # Extract top N bigrams
        top_bigrams = bigrams[:top_n]
        
        # Create bar chart
        if not top_bigrams:
            return None
            
        bigram_df = pd.DataFrame(top_bigrams, columns=['Bigram', 'Count'])
        
        fig = px.bar(
            bigram_df, 
            y='Bigram', 
            x='Count', 
            orientation='h',
            title=f'Top {top_n} Bigrams in {column}',
            height=500,
            color='Count',
            color_continuous_scale=px.colors.sequential.Blues
        )
        
        fig.update_layout(
            xaxis_title='Frequency',
            yaxis_title='Bigram',
            yaxis={'categoryorder':'total ascending'},
            template="plotly_dark",
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(30,30,30,1)'
        )
        
        return fig
    except Exception as e:
        logger.error(f"Error in bigram analysis: {str(e)}")
        return None

def compare_abstract_topics(df_before, df_after):
    """Compare abstract topics between before and after periods"""
    if len(df_before) == 0 or len(df_after) == 0 or 'Abstract' not in df_before.columns or 'Abstract' not in df_after.columns:
        return None
    
    try:
        # Get abstracts
        texts_before = df_before['Abstract'].dropna().astype(str).tolist()
        texts_after = df_after['Abstract'].dropna().astype(str).tolist()
        
        if not texts_before or not texts_after:
            return None
        
        # Use English stopwords (as string)
        vectorizer = TfidfVectorizer(
            stop_words="english",
            max_features=1000,
            min_df=2
        )
        
        # Get combined TF-IDF
        all_texts = texts_before + texts_after
        vectorizer.fit(all_texts)
        
        # Transform each corpus
        X_before = vectorizer.transform(texts_before)
        X_after = vectorizer.transform(texts_after)
        
        # Get average TF-IDF for each term in each corpus
        avg_tfidf_before = np.asarray(X_before.mean(axis=0)).flatten()
        avg_tfidf_after = np.asarray(X_after.mean(axis=0)).flatten()
        
        # Get feature names
        feature_names = vectorizer.get_feature_names_out()
        
        # Find differential terms (terms with largest difference in TF-IDF)
        diff_scores = []
        for i, term in enumerate(feature_names):
            diff = avg_tfidf_after[i] - avg_tfidf_before[i]
            diff_scores.append((term, diff))
        
        # Sort by absolute difference
        diff_scores.sort(key=lambda x: abs(x[1]), reverse=True)
        
        # Get top differential terms
        top_n = 20
        top_diff = diff_scores[:top_n]
        
        # Create dataframe for visualization
        diff_df = pd.DataFrame(top_diff, columns=['Term', 'Difference'])
        diff_df['Period'] = ['After Approval' if d > 0 else 'Before Approval' for d in diff_df['Difference']]
        diff_df['Abs_Difference'] = diff_df['Difference'].abs()
        
        # Create horizontal bar chart
        fig = px.bar(
            diff_df, 
            y='Term', 
            x='Difference',
            orientation='h',
            color='Period',
            title=f'Top {top_n} Differential Terms Between Before and After Approval',
            height=600,
            color_discrete_map={
                'Before Approval': '#E53935',
                'After Approval': '#43A047'
            }
        )
        
        fig.update_layout(
            xaxis_title='TF-IDF Difference (After - Before)',
            yaxis_title='Term',
            yaxis={'categoryorder':'total ascending'},
            template="plotly_dark",
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(30,30,30,1)'
        )
        
        return fig
    except Exception as e:
        logger.error(f"Error in topic comparison: {str(e)}")
        return None

def journal_distribution_pie(df):
    """Create a pie chart of journal distribution"""
    if len(df) == 0 or 'JournalTitle' not in df.columns:
        return None
    
    try:
        # Get journal counts
        journal_counts = df['JournalTitle'].value_counts().reset_index()
        journal_counts.columns = ['Journal', 'Count']
        
        # Keep top 8 journals, group others
        top_n = min(8, len(journal_counts))
        if len(journal_counts) > top_n:
            top_journals = journal_counts.head(top_n)
            other_count = journal_counts.iloc[top_n:]['Count'].sum()
            top_journals = pd.concat([
                top_journals, 
                pd.DataFrame({'Journal': ['Other Journals'], 'Count': [other_count]})
            ])
        else:
            top_journals = journal_counts
        
        # Create pie chart
        fig = px.pie(
            top_journals, 
            values='Count', 
            names='Journal',
            title='Distribution of Publications by Journal',
            hole=0.4,
            color_discrete_sequence=px.colors.qualitative.Set2
        )
        
        fig.update_traces(textposition='inside', textinfo='percent+label')
        fig.update_layout(
            uniformtext_minsize=10, 
            uniformtext_mode='hide',
            template="plotly_dark",
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    except Exception as e:
        logger.error(f"Error creating journal pie chart: {str(e)}")
        return None

def extract_statistical_insights(df_before, df_after):
    """Extract key statistical insights from the datasets"""
    insights = []
    
    try:
        # 1. Publication volume
        before_count = len(df_before)
        after_count = len(df_after)
        
        if before_count > 0 and after_count > 0:
            percent_change = ((after_count - before_count) / before_count) * 100
            change_direction = "increase" if percent_change > 0 else "decrease"
            
            insights.append(f"Publication volume showed a {abs(percent_change):.1f}% {change_direction} after FDA approval ({after_count} vs. {before_count} publications).")
        
        # 2. Journal diversity
        if 'JournalTitle' in df_before.columns and 'JournalTitle' in df_after.columns:
            unique_journals_before = df_before['JournalTitle'].nunique()
            unique_journals_after = df_after['JournalTitle'].nunique()
            
            if unique_journals_before > 0 and unique_journals_after > 0:
                journal_percent_change = ((unique_journals_after - unique_journals_before) / unique_journals_before) * 100
                journal_change_direction = "more diverse" if journal_percent_change > 0 else "less diverse"
                
                insights.append(f"Research became {journal_change_direction} after approval with {unique_journals_after} journals compared to {unique_journals_before} before.")
        
        # 3. Author count
        author_cols_before = [col for col in df_before.columns if col.startswith('Author')]
        author_cols_after = [col for col in df_after.columns if col.startswith('Author')]
        
        avg_authors_before = 0
        avg_authors_after = a = 0
        
        if author_cols_before:
            # Count non-empty author cells per row
            authors_per_paper_before = df_before[author_cols_before].notna().sum(axis=1)
            avg_authors_before = authors_per_paper_before.mean()
        
        if author_cols_after:
            authors_per_paper_after = df_after[author_cols_after].notna().sum(axis=1)
            avg_authors_after = authors_per_paper_after.mean()
        
        if avg_authors_before > 0 and avg_authors_after > 0:
            author_change = avg_authors_after - avg_authors_before
            author_direction = "more" if author_change > 0 else "fewer"
            
            insights.append(f"Research papers had {author_direction} authors on average after approval ({avg_authors_after:.1f} vs. {avg_authors_before:.1f} authors per paper).")
        
        # 4. Abstract length and complexity
        if 'Abstract' in df_before.columns and 'Abstract' in df_after.columns:
            # Average abstract length
            abstract_length_before = df_before['Abstract'].str.len().mean()
            abstract_length_after = df_after['Abstract'].str.len().mean()
            
            if not pd.isna(abstract_length_before) and not pd.isna(abstract_length_after):
                length_change = ((abstract_length_after - abstract_length_before) / abstract_length_before) * 100
                length_direction = "longer" if length_change > 0 else "shorter"
                
                insights.append(f"Abstracts became {abs(length_change):.1f}% {length_direction} after approval ({abstract_length_after:.0f} vs. {abstract_length_before:.0f} characters on average).")
        
        # 5. Publication year span
        if 'PublicationDate' in df_before.columns and 'PublicationDate' in df_after.columns:
            # Extract years from the full date string
            df_before_year = df_before.copy()
            df_after_year = df_after.copy()
            
            df_before_year['Year'] = df_before_year['PublicationDate'].str[:4].astype(float, errors='ignore')
            df_after_year['Year'] = df_after_year['PublicationDate'].str[:4].astype(float, errors='ignore')
            
            # Get year ranges
            before_min_year = df_before_year['Year'].min()
            before_max_year = df_before_year['Year'].max()
            after_min_year = df_after_year['Year'].min()
            after_max_year = df_after_year['Year'].max()
            
            if not pd.isna(before_min_year) and not pd.isna(before_max_year) and not pd.isna(after_min_year) and not pd.isna(after_max_year):
                insights.append(f"Before approval research spans {before_max_year - before_min_year + 1:.0f} years ({before_min_year:.0f}-{before_max_year:.0f}), while after approval spans {after_max_year - after_min_year + 1:.0f} years ({after_min_year:.0f}-{after_max_year:.0f}).")
        
    except Exception as e:
        logger.error(f"Error generating insights: {str(e)}")
        
    return insights

# ------------------ File Conversion Functions ------------------

def create_hierarchical_excel(df):
    """
    Create a hierarchical Excel file with authors as sub-rows.
    
    The Excel will have:
    - Main rows for article details
    - Sub-rows for each author and their affiliation
    - Merged cells for article details across all sub-rows
    
    Returns the Excel file as bytes.
    """
    if is_dataframe_empty(df):
        # Return an empty Excel file with just headers if df is empty
        output = BytesIO()
        pd.DataFrame(columns=['No Data Available']).to_excel(output, index=False)
        return output.getvalue()
    
    # Identify non-author columns and author-related columns
    author_pattern = re.compile(r"^Author(\d+)$")
    affiliation_pattern = re.compile(r"^Affiliation(\d+)$")
    
    article_columns = []
    for col in df.columns:
        if not author_pattern.match(col) and not affiliation_pattern.match(col):
            article_columns.append(col)
    
    # Get all author-affiliation pairs
    author_data = []
    for _, row in df.iterrows():
        article_info = {col: row[col] for col in article_columns}
        
        authors = []
        for i in range(1, 100):  # Assuming no more than 100 authors
            author_col = f"Author{i}"
            affiliation_col = f"Affiliation{i}"
            
            if author_col not in df.columns:
                break
                
            if pd.notna(row.get(author_col)) and row.get(author_col) != '':
                author_info = {
                    'Author': row.get(author_col, ''),
                    'Affiliation': row.get(affiliation_col, '')
                }
                authors.append(author_info)
        
        if authors:
            author_data.append({
                'article': article_info,
                'authors': authors
            })
    
    # Create a new dataframe with the hierarchical structure
    rows = []
    
    for article in author_data:
        article_info = article['article']
        authors = article['authors']
        
        for i, author in enumerate(authors):
            row = {}
            
            # Only include article info in the first author row
            if i == 0:
                for col in article_columns:
                    row[col] = article_info.get(col, '')
            else:
                for col in article_columns:
                    row[col] = ''  # Empty for merged cells
            
            row['Author'] = author['Author']
            row['Affiliation'] = author['Affiliation']
            rows.append(row)
    
    # Create a new DataFrame with the hierarchical structure
    columns = article_columns + ['Author', 'Affiliation']
    hierarchical_df = pd.DataFrame(rows, columns=columns)
    
    # Create Excel in memory
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    hierarchical_df.to_excel(writer, index=False, sheet_name='Articles')
    
    # Get the worksheet
    worksheet = writer.sheets['Articles']
    
    # Define styles
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    wrapped_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Side(style='thin')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Format header row
    for col_idx, col_name in enumerate(hierarchical_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border
    
    # Track where each article starts and ends
    article_ranges = []
    current_start = 2  # Start from row 2 (after header)
    current_pmid = None
    
    for row_idx, row in enumerate(hierarchical_df.iterrows(), 2):  # Start from row 2
        row_data = row[1]  # Get the row data
        pmid = row_data.get('PMID', '')
        
        if pmid and pmid != current_pmid:
            if current_pmid is not None:
                article_ranges.append((current_start, row_idx - 1))
            current_start = row_idx
            current_pmid = pmid
    
    # Add the last article range
    if current_pmid is not None:
        article_ranges.append((current_start, row_idx))
    
    # Apply alternating row colors for each article group
    is_alt_row = False
    for start_row, end_row in article_ranges:
        if is_alt_row:
            for row in range(start_row, end_row + 1):
                for col in range(1, len(hierarchical_df.columns) + 1):
                    worksheet.cell(row=row, column=col).fill = alt_row_fill
        is_alt_row = not is_alt_row
    
    # Merge cells for each article
    for start_row, end_row in article_ranges:
        if start_row == end_row:
            continue  # No need to merge for single-author articles
        
        for col_idx, col_name in enumerate(article_columns, 1):
            # Get column letter
            col_letter = get_column_letter(col_idx)
            # Merge cells
            worksheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
            # Center the content vertically
            cell = worksheet.cell(row=start_row, column=col_idx)
            cell.alignment = wrapped_alignment
    
    # Add borders to all cells
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).border = border
    
    # Adjust column widths
    for col_idx, col_name in enumerate(hierarchical_df.columns, 1):
        column_width = 15  # Default width
        
        if col_name == 'Abstract':
            column_width = 50
        elif col_name == 'ArticleTitle':
            column_width = 40
        elif col_name == 'JournalTitle':
            column_width = 30
        elif col_name == 'Affiliation':
            column_width = 40
        elif col_name == 'PublicationDate':
            column_width = 15
        elif col_name == 'PMID':
            column_width = 12
        elif col_name == 'Author':
            column_width = 25
        elif col_name == 'DOI':
            column_width = 30
        elif col_name == 'ArticleLink':
            column_width = 25
        
        worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
    
    # Make the ArticleLink column clickable
    if 'ArticleLink' in hierarchical_df.columns:
        article_link_idx = hierarchical_df.columns.get_loc('ArticleLink') + 1  # +1 because Excel is 1-indexed
        for row_idx, row in enumerate(hierarchical_df.iterrows(), 2):  # Start from row 2
            row_data = row[1]
            link = row_data.get('ArticleLink', '')
            if link:
                cell = worksheet.cell(row=row_idx, column=article_link_idx)
                cell.hyperlink = link
                cell.value = "PubMed Link"
                cell.font = Font(color="0000FF", underline="single")
    
    # Freeze the top row
    worksheet.freeze_panes = "A2"
    
    # Save and return
    writer.close()
    output.seek(0)
    return output.getvalue()

def to_excel(df):
    """Convert dataframe to Excel file in memory"""
    try:
        if is_dataframe_empty(df):
            # Return an empty Excel file with just headers if df is empty
            output = BytesIO()
            pd.DataFrame(columns=['No Data Available']).to_excel(output, index=False)
            return output.getvalue()
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Auto-adjust columns' width and make ArticleLink clickable
            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(df.columns):
                # Set column width
                column_width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width + 2  # Add padding
                
                # Make ArticleLink column clickable with hyperlinks
                if col == 'ArticleLink':
                    for row_idx, link in enumerate(df[col], 2):  # Start from row 2 (after header)
                        if pd.notna(link) and link != '':
                            cell = worksheet.cell(row=row_idx, column=i+1)
                            cell.hyperlink = link
                            cell.value = "PubMed Link"
                            cell.font = Font(color="0000FF", underline="single")
        
        return output.getvalue()
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        # Create a simple Excel file without formatting as fallback
        output = BytesIO()
        df.to_excel(output, index=False)
        return output.getvalue()

# ------------------ Main Streamlit App ------------------

# Initialize session state
if 'df_before' not in st.session_state:
    st.session_state.df_before = None
if 'df_after' not in st.session_state:
    st.session_state.df_after = None
if 'stats_before' not in st.session_state:
    st.session_state.stats_before = None
if 'stats_after' not in st.session_state:
    st.session_state.stats_after = None
if 'insights' not in st.session_state:
    st.session_state.insights = None
if 'excel_before' not in st.session_state:
    st.session_state.excel_before = None
if 'excel_after' not in st.session_state:
    st.session_state.excel_after = None
if 'hierarchical_excel_before' not in st.session_state:
    st.session_state.hierarchical_excel_before = None
if 'hierarchical_excel_after' not in st.session_state:
    st.session_state.hierarchical_excel_after = None
if 'search_submitted' not in st.session_state:
    st.session_state.search_submitted = False
if 'moved_to_before' not in st.session_state:
    st.session_state.moved_to_before = 0
if 'moved_to_after' not in st.session_state:
    st.session_state.moved_to_after = 0
# Add to the session state initialization section
if 'deduped_before' not in st.session_state:
    st.session_state.deduped_before = 0
if 'deduped_after' not in st.session_state:
    st.session_state.deduped_after = 0
    
    # Helper function to safely check if DataFrame is empty
def is_dataframe_empty(df):
    if df is None:
        return True
    if not isinstance(df, pd.DataFrame):
        return True
    return df.empty

# Sidebar configuration
st.sidebar.markdown("### Search Configuration")

with st.sidebar.form("search_form"):
    drug_name = st.text_input("Drug Name", value="Quviviq")
    composition = st.text_input("Composition/Compound", value="daridorexant")
    disease = st.text_input("Target Disease", value="Insomnia")
    company = st.text_input("Company", value="Idorsia Pharmaceuticals US Inc")
    # Set a wide date range to allow for historical FDA approvals (1950-2030)
    fda_approval_date = st.date_input(
        "FDA Approval Date", 
        value=datetime.date(2022, 1, 7),
        min_value=datetime.date(1950, 1, 1),
        max_value=datetime.date(2030, 12, 31)
    )
    
    # Add flexibility for date ranges
    st.write("##### Date Range Configuration")
    col1, col2 = st.columns(2)
    with col1:
        before_years = st.number_input("Years before FDA approval", min_value=1, max_value=50, value=10)
    with col2:
        after_years = st.number_input("Years after FDA approval (max to current)", min_value=1, max_value=50, value=50)
    
    include_company = st.checkbox("Include company in search query", value=False)
    article_limit = st.number_input("Limit articles (0 for no limit)", min_value=0, value=0)
    use_improved_search = st.checkbox("Use improved search strategy", value=False, help="Include both drug name and compound in both before/after queries")
    
    submitted = st.form_submit_button("Search PubMed")
    
    if submitted:
        st.session_state.search_submitted = True

# Process when form is submitted
if submitted or st.session_state.search_submitted:
    with st.spinner("Fetching articles from PubMed..."):
        # Convert dates with flexible range - with safer date arithmetic
        fda_date = datetime.datetime.combine(fda_approval_date, datetime.datetime.min.time())
        
        # Helper function to safely subtract years from a date
        def subtract_years(dt, years):
            try:
                return dt.replace(year=dt.year - years)
            except ValueError:
                # Handle Feb 29 and similar edge cases by moving to the last day of the month
                # For example, Feb 29, 2020 - 1 year becomes Feb 28, 2019
                new_year = dt.year - years
                new_month = dt.month
                # Find the last day of the month in the new year
                last_day = 28  # Default for February
                if new_month in [1, 3, 5, 7, 8, 10, 12]:
                    last_day = 31
                elif new_month in [4, 6, 9, 11]:
                    last_day = 30
                elif new_month == 2 and (new_year % 4 == 0 and (new_year % 100 != 0 or new_year % 400 == 0)):
                    last_day = 29  # Leap year
                
                return dt.replace(year=new_year, day=min(dt.day, last_day))
        
        # Helper function to safely add years to a date
        def add_years(dt, years):
            try:
                return dt.replace(year=dt.year + years)
            except ValueError:
                # Handle Feb 29 and similar edge cases
                new_year = dt.year + years
                new_month = dt.month
                # Find the last day of the month in the new year
                last_day = 28  # Default for February
                if new_month in [1, 3, 5, 7, 8, 10, 12]:
                    last_day = 31
                elif new_month in [4, 6, 9, 11]:
                    last_day = 30
                elif new_month == 2 and (new_year % 4 == 0 and (new_year % 100 != 0 or new_year % 400 == 0)):
                    last_day = 29  # Leap year
                
                return dt.replace(year=new_year, day=min(dt.day, last_day))
        
        before_start_date = subtract_years(fda_date, before_years)
        before_end_date = fda_date
        after_start_date = fda_date
        after_end_date = min(datetime.datetime.today(), add_years(fda_date, after_years))
        
        # Format dates for PubMed query
        before_start_date_str = before_start_date.strftime("%Y/%m/%d")
        before_end_date_str = before_end_date.strftime("%Y/%m/%d")
        after_start_date_str = after_start_date.strftime("%Y/%m/%d")
        after_end_date_str = after_end_date.strftime("%Y/%m/%d")
        
        # Build queries using improved strategy if selected
        if use_improved_search:
            # Use both drug name and composition in both queries for better coverage
            query_before = (f'("{composition}"[All Fields] OR "{drug_name}"[All Fields]) AND "{disease}"[All Fields] '
                         f'AND ("{before_start_date_str}"[pdat] : "{before_end_date_str}"[pdat])')
            
            query_after = (f'("{drug_name}"[All Fields] OR "{composition}"[All Fields]) '
                         f'AND "{disease}"[All Fields] ')
        else:
            # Original simpler query strategy
            query_before = (f'"{composition}"[All Fields] AND "{disease}"[All Fields] '
                         f'AND ("{before_start_date_str}"[pdat] : "{before_end_date_str}"[pdat])')
            
            query_after = (f'("{drug_name}"[All Fields] OR "{composition}"[All Fields]) '
                         f'AND "{disease}"[All Fields] ')
        
        # Add company to query if requested
        if include_company:
            query_after += f'AND "{company}"[All Fields] '
            
        query_after += f'AND ("{after_start_date_str}"[pdat] : "{after_end_date_str}"[pdat])'
        
        st.subheader("Search Queries")
        st.code(f"Before Approval: {query_before}")
        st.code(f"After Approval: {query_after}")
        
        # Only fetch if we need to (if data not already in session state)
        if st.session_state.df_before is None or submitted:
            # Fetch articles
            st.write("Fetching articles for Before Approval period...")
            progress_bar_before = st.progress(0)
            limit_value = int(article_limit) if article_limit > 0 else None
            records_before = fetch_pubmed_articles(
                query_before, 
                limit=limit_value,
                progress_callback=lambda progress: progress_bar_before.progress(progress)
            )
            
            st.write("Fetching articles for After Approval period...")
            progress_bar_after = st.progress(0)
            records_after = fetch_pubmed_articles(
                query_after, 
                limit=limit_value,
                progress_callback=lambda progress: progress_bar_after.progress(progress)
            )
            
            # Verify dates and recategorize records for better precision
            verified_before, verified_after, moved_to_before, moved_to_after = verify_dates_and_recategorize(
                records_before, records_after, fda_date
            )

            # Deduplicate records to ensure no overlap
            final_before, final_after = deduplicate_records(verified_before, verified_after)

            # Calculate how many were deduped
            before_deduped = len(verified_before) - len(final_before)
            after_deduped = len(verified_after) - len(final_after)

            st.session_state.moved_to_before = moved_to_before
            st.session_state.moved_to_after = moved_to_after
            st.session_state.deduped_before = before_deduped
            st.session_state.deduped_after = after_deduped

            # Convert to dataframes
            df_before_raw = pd.DataFrame(final_before)
            df_after_raw = pd.DataFrame(final_after)
            
            # Process and verify data
            st.subheader("Processing and Validating Data")
            
            if moved_to_before > 0 or moved_to_after > 0:
                st.info(f"Date verification: {moved_to_after} articles moved to 'after approval' and {moved_to_before} articles moved to 'before approval' based on precise publication dates.")

            if before_deduped > 0 or after_deduped > 0:
                st.info(f"Deduplication: Removed {before_deduped} duplicate articles from 'before approval' and {after_deduped} duplicate articles from 'after approval' datasets.")
                        
            if len(df_before_raw) > 0:
                st.write(f"Processing {len(df_before_raw)} articles from Before Approval period...")
                df_before, stats_before = process_dataset(df_before_raw, "before", before_start_date, before_end_date)
                st.session_state.df_before = df_before
                st.session_state.stats_before = stats_before
                st.session_state.excel_before = to_excel(df_before)
                st.session_state.hierarchical_excel_before = create_hierarchical_excel(df_before)
            else:
                st.info("No articles found for Before Approval period.")
                st.session_state.df_before = pd.DataFrame()
                st.session_state.stats_before = {"original_rows": 0}
                st.session_state.excel_before = to_excel(pd.DataFrame())
                st.session_state.hierarchical_excel_before = create_hierarchical_excel(pd.DataFrame())
            
            if len(df_after_raw) > 0:
                st.write(f"Processing {len(df_after_raw)} articles from After Approval period...")
                df_after, stats_after = process_dataset(df_after_raw, "after", after_start_date, after_end_date)
                st.session_state.df_after = df_after
                st.session_state.stats_after = stats_after
                st.session_state.excel_after = to_excel(df_after)
                st.session_state.hierarchical_excel_after = create_hierarchical_excel(df_after)
            else:
                st.info("No articles found for After Approval period.")
                st.session_state.df_after = pd.DataFrame()
                st.session_state.stats_after = {"original_rows": 0}
                st.session_state.excel_after = to_excel(pd.DataFrame())
                st.session_state.hierarchical_excel_after = create_hierarchical_excel(pd.DataFrame())
            
            # Generate insights
            st.session_state.insights = extract_statistical_insights(
                st.session_state.df_before, 
                st.session_state.df_after
            )
            
            st.success("Processing complete! You can now explore the analysis tabs below.")
        
        # Display tabs for results
        if st.session_state.df_before is not None or st.session_state.df_after is not None:
            tabs = st.tabs([
                "📊 Data Overview", 
                "📈 Time Trends", 
                "🔎 Content Analysis", 
                "👥 Author Analysis",
                "💾 Download Data",
                "ℹ️ About"
            ])
            
            # Tab 1: Data Overview
            with tabs[0]:
                st.markdown("<h2 class='subheader'>Data Overview</h2>", unsafe_allow_html=True)
                
                # Display processing statistics
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("<h3 class='section-header'>Before Approval</h3>", unsafe_allow_html=True)
                    if st.session_state.stats_before:
                        stats = st.session_state.stats_before
                        st.write(f"Articles found: {stats.get('original_rows', 0)}")
                        if stats.get('original_rows', 0) > 0:
                            st.write(f"Articles after date filtering: {stats.get('final_rows', 0)}")
                            st.write(f"Invalid dates removed: {stats.get('invalid_dates', 0)}")
                            st.write(f"Maximum author number: {stats.get('max_author', 0)}")
                            st.write(f"Columns trimmed: {stats.get('columns_removed', 0)}")
                        
                        # Preview the data
                        if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                            st.write("Data Preview:")
                            st.dataframe(st.session_state.df_before.head(5), use_container_width=True)
                
                with col2:
                    st.markdown("<h3 class='section-header'>After Approval</h3>", unsafe_allow_html=True)
                    if st.session_state.stats_after:
                        stats = st.session_state.stats_after
                        st.write(f"Articles found: {stats.get('original_rows', 0)}")
                        if stats.get('original_rows', 0) > 0:
                            st.write(f"Articles after date filtering: {stats.get('final_rows', 0)}")
                            st.write(f"Invalid dates removed: {stats.get('invalid_dates', 0)}")
                            st.write(f"Maximum author number: {stats.get('max_author', 0)}")
                            st.write(f"Columns trimmed: {stats.get('columns_removed', 0)}")
                        
                        # Preview the data
                        if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                            st.write("Data Preview:")
                            st.dataframe(st.session_state.df_after.head(5), use_container_width=True)
                
                # Display key insights
                st.markdown("<h3 class='section-header'>Key Insights</h3>", unsafe_allow_html=True)
                
                if st.session_state.insights and len(st.session_state.insights) > 0:
                    for insight in st.session_state.insights:
                        st.markdown(f"<div class='insight-box'>💡 {insight}</div>", unsafe_allow_html=True)
                else:
                    st.info("No significant insights could be generated from the data. Try adjusting your search parameters or including more articles.")
                
                # Journal distribution visualizations
                st.markdown("<h3 class='section-header'>Journal Distribution</h3>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("Before Approval:")
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        chart = create_top_journals_chart(st.session_state.df_before)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient journal data for visualization.")
                    else:
                        st.info("No data available for visualization.")
                
                with col2:
                    st.write("After Approval:")
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        chart = create_top_journals_chart(st.session_state.df_after)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient journal data for visualization.")
                    else:
                        st.info("No data available for visualization.")
            
            # Tab 2: Time Trends
            with tabs[1]:
                st.markdown("<h2 class='subheader'>Publication Trends Over Time</h2>", unsafe_allow_html=True)
                
                # Publication trends over time
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("<h3 class='section-header'>Before Approval Trend</h3>", unsafe_allow_html=True)
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        chart = create_yearly_trend_chart(st.session_state.df_before)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient year data for trend visualization.")
                        
                        # Add monthly trend chart if data is available
                        monthly_chart = create_monthly_trend_chart(st.session_state.df_before)
                        if monthly_chart:
                            st.plotly_chart(monthly_chart, use_container_width=True)
                    else:
                        st.info("No data available for trend visualization.")
                
                with col2:
                    st.markdown("<h3 class='section-header'>After Approval Trend</h3>", unsafe_allow_html=True)
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        chart = create_yearly_trend_chart(st.session_state.df_after)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient year data for trend visualization.")
                        
                        # Add monthly trend chart if data is available
                        monthly_chart = create_monthly_trend_chart(st.session_state.df_after)
                        if monthly_chart:
                            st.plotly_chart(monthly_chart, use_container_width=True)
                    else:
                        st.info("No data available for trend visualization.")
                
                # Journal distribution pie charts
                st.markdown("<h3 class='section-header'>Journal Distribution Over Time</h3>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("Before Approval:")
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        chart = journal_distribution_pie(st.session_state.df_before)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient journal data for visualization.")
                    else:
                        st.info("No data available for visualization.")
                
                with col2:
                    st.write("After Approval:")
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        chart = journal_distribution_pie(st.session_state.df_after)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient journal data for visualization.")
                    else:
                        st.info("No data available for visualization.")
            
            # Tab 3: Content Analysis
            with tabs[2]:
                st.markdown("<h2 class='subheader'>Content Analysis</h2>", unsafe_allow_html=True)
                
                # Word clouds
                st.markdown("<h3 class='section-header'>Abstract Word Clouds</h3>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                
                # In the Content Analysis tab where wordcloud figures are displayed:
                with col1:
                    st.write("Before Approval:")
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        fig = generate_wordcloud(st.session_state.df_before, column='Abstract')
                        if fig:
                            # Use st.pyplot() with explicit figure
                            st.pyplot(fig)
                        else:
                            st.info("Insufficient abstract data for word cloud.")
                    else:
                        st.info("No data available for word cloud.")
                
                with col2:
                    st.write("After Approval:")
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        fig = generate_wordcloud(st.session_state.df_after, column='Abstract')
                        if fig:
                            st.pyplot(fig)
                        else:
                            st.info("Insufficient abstract data for word cloud.")
                    else:
                        st.info("No data available for word cloud.")
                
                # Bigram analysis
                st.markdown("<h3 class='section-header'>Bigram Analysis</h3>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("Before Approval:")
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        chart = generate_bigram_analysis(st.session_state.df_before, column='Abstract')
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient abstract data for bigram analysis.")
                    else:
                        st.info("No data available for bigram analysis.")
                
                with col2:
                    st.write("After Approval:")
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        chart = generate_bigram_analysis(st.session_state.df_after, column='Abstract')
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient abstract data for bigram analysis.")
                    else:
                        st.info("No data available for bigram analysis.")
                
                # Topic comparison
                st.markdown("<h3 class='section-header'>Topic Comparison</h3>", unsafe_allow_html=True)
                
                if st.session_state.df_before is not None and st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_before) and not is_dataframe_empty(st.session_state.df_after):
                    chart = compare_abstract_topics(st.session_state.df_before, st.session_state.df_after)
                    if chart:
                        st.plotly_chart(chart, use_container_width=True)
                    else:
                        st.info("Insufficient data for topic comparison.")
                else:
                    st.info("Data from both before and after approval periods is required for topic comparison.")
            
            # Tab 4: Author Analysis
            with tabs[3]:
                st.markdown("<h2 class='subheader'>Author Analysis</h2>", unsafe_allow_html=True)
                
                # Author network
                st.markdown("<h3 class='section-header'>Author Collaboration Networks</h3>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("Before Approval:")
                    if st.session_state.df_before is not None and not is_dataframe_empty(st.session_state.df_before):
                        top_n = st.slider("Number of top authors (Before)", min_value=5, max_value=50, value=15, key="before_author_slider")
                        chart = create_author_collaboration_network(st.session_state.df_before, top_n=top_n)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient author data for collaboration network.")
                    else:
                        st.info("No data available for author analysis.")
                
                with col2:
                    st.write("After Approval:")
                    if st.session_state.df_after is not None and not is_dataframe_empty(st.session_state.df_after):
                        top_n = st.slider("Number of top authors (After)", min_value=5, max_value=50, value=15, key="after_author_slider")
                        chart = create_author_collaboration_network(st.session_state.df_after, top_n=top_n)
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
                        else:
                            st.info("Insufficient author data for collaboration network.")
                    else:
                        st.info("No data available for author analysis.")
            
            # Tab 5: Download Data
            with tabs[4]:
                st.markdown("<h2 class='subheader'>Download Processed Data</h2>", unsafe_allow_html=True)
                
                st.write("""
                The processed data files below have been cleaned and optimized:
                - Verified dates are within appropriate ranges
                - Removed excess empty columns
                - Formatted for easy analysis
                - DOI information included in a separate column
                - Clickable PubMed links included
                """)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("<h3 class='section-header'>Before Approval Data</h3>", unsafe_allow_html=True)
                    if st.session_state.excel_before is not None:
                        if st.session_state.df_before is None or is_dataframe_empty(st.session_state.df_before):
                            st.info("No data available for download.")
                        else:
                            col1a, col1b = st.columns(2)
                            with col1a:
                                download_before = st.download_button(
                                    label="Download Standard Excel",
                                    data=st.session_state.excel_before,
                                    file_name=f"{drug_name}_BeforeApproval.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Download the processed data in standard Excel format"
                                )
                            with col1b:
                                download_hierarchical = st.download_button(
                                    label="Download Hierarchical Excel",
                                    data=st.session_state.hierarchical_excel_before,
                                    file_name=f"{drug_name}_BeforeApproval_Hierarchical.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Download in hierarchical format with authors as sub-rows"
                                )
                            st.write(f"Contains {len(st.session_state.df_before)} articles")
                    else:
                        st.info("Data not yet processed. Run a search first.")
                
                with col2:
                    st.markdown("<h3 class='section-header'>After Approval Data</h3>", unsafe_allow_html=True)
                    if st.session_state.excel_after is not None:
                        if st.session_state.df_after is None or is_dataframe_empty(st.session_state.df_after):
                            st.info("No data available for download.")
                        else:
                            col2a, col2b = st.columns(2)
                            with col2a:
                                download_after = st.download_button(
                                    label="Download Standard Excel",
                                    data=st.session_state.excel_after,
                                    file_name=f"{drug_name}_AfterApproval.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Download the processed data in standard Excel format",
                                    key="download_after_button"
                                )
                            with col2b:
                                download_hierarchical = st.download_button(
                                    label="Download Hierarchical Excel",
                                    data=st.session_state.hierarchical_excel_after,
                                    file_name=f"{drug_name}_AfterApproval_Hierarchical.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    help="Download in hierarchical format with authors as sub-rows",
                                    key="download_hierarchical_after_button"
                                )
                            st.write(f"Contains {len(st.session_state.df_after)} articles")
                    else:
                        st.info("Data not yet processed. Run a search first.")
            
            # Tab 6: About
            with tabs[5]:
                st.markdown("<h2 class='subheader'>About the PubMed Research Analyzer</h2>", unsafe_allow_html=True)
                
                st.markdown("""
                ### Overview
                
                The PubMed Research Analyzer is a tool designed to help researchers analyze publication trends before and after FDA approval of pharmaceutical products. It provides insights into how research patterns, journal distributions, author collaborations, and content focus change following regulatory approval.
                
                ### Key Features
                
                - **Precision Date Handling**: Uses month-level precision to accurately categorize publications
                - **Flexible Date Ranges**: Customizable time periods before and after FDA approval
                - **Smart Search Strategy**: Includes both brand name and compound name for comprehensive results
                - **DOI Information**: Includes DOI for each article when available
                - **Clickable Links**: Direct access to articles on PubMed for easy reference
                - **Hierarchical Excel Export**: Creates beautifully formatted Excel files with authors as sub-rows
                - **Date Verification**: Double-checks publication dates to ensure correct categorization
                - **Visual Analysis**: Generates visualizations of publication trends, journal distributions, and content analysis
                - **Author Network Mapping**: Creates collaboration networks showing key researchers in the field
                
                ### How to Use
                
                1. Enter the drug information in the sidebar
                2. Set the FDA approval date
                3. Configure the date ranges for before and after approval periods
                4. Click "Search PubMed" to retrieve articles
                5. Explore the analysis in the different tabs
                6. Download data in standard or hierarchical Excel format with DOIs and clickable links
                
                ### About Hierarchical Excel Format
                
                The hierarchical Excel format organizes data with:
                - Main rows containing article information (including DOI and clickable PubMed links)
                - Sub-rows listing each author and their affiliation
                - Merged cells for article details
                - Alternating row colors for better readability
                - Appropriate column widths and text wrapping
                
                This format makes it easier to explore author contributions while maintaining the article context.
                """)
                
                st.markdown("### Data Processing")
                st.image("https://www.ncbi.nlm.nih.gov/core/assets/cbe/images/logo-pubmed.svg", width=200)
                st.markdown("""
                All publication data is retrieved from PubMed using the E-utilities API. The application processes XML responses to extract:
                
                - Publication metadata (title, journal, dates)
                - DOIs and article links
                - Author information
                - Abstract content
                
                The data is then processed for analysis, including date normalization, author extraction, and statistical calculations.
                """)

# Footer
st.markdown("""
<div class="footer">
    <p>Developed by Rhenix Life Sciences</p>
    <p>© 2025 Rhenix Life Sciences. All rights reserved.</p>
    <p>Data sourced from PubMed via E-utilities API</p>
</div>
""", unsafe_allow_html=True)